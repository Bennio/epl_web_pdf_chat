import os
import openpyxl
import pandas as pd
import requests
from dotenv import load_dotenv

load_dotenv()

# chemin vers le fichier Excel
# excel_file_path = '/Users/bendylatortue/Documents/Data_Evidence.xlsx'
excel_file_path = '/Users/bendylatortue/PycharmProjects/chatWithWebsiteAndPDF/Process/Updated_Data_Evidence.xlsx'
api_endpoint = 'http://localhost:8000/backend/api/conversations/process-document/'
auth_endpoint = 'http://localhost:8000/login/'

username = os.environ.get('CHATBOT_APP_USERNAME')
password = os.environ.get('CHATBOT_APP_PASSWORD')

response = requests.post(auth_endpoint, json={"username": username, "password": password})

# Vérifier si la requête a réussi
if response.status_code == 200:
    # Extraire le jeton du corps de la réponse
    token = response.json()['token']
    print(f"Token obtained: {token}")
else:
    print(f"Failed to obtain token. Status code: {response.status_code}")

headers = {
    'Authorization': f'Token {token}'
}


def send_to_process_document(url, questions):
    response = requests.post(api_endpoint, json={'url': url, 'questions': questions}, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Failed to send data to the API. Status code: {response.status_code}")
        return None


# Fonction pour obtenir les en-têtes des cellules fusionnées
def get_merged_cell_headers(sheet):
    headers = []
    # Récupérer les plages de cellules fusionnées dans les deux premières lignes
    merged_cells_ranges = list(sheet.merged_cells.ranges)

    # Construire un dictionnaire pour garder la trace des titres d'en-têtes pour les colonnes fusionnées
    merged_headers = {}
    for range_ in merged_cells_ranges:
        # On prend l'en-tête du haut gauche de la cellule fusionnée
        top_left_cell_value = sheet.cell(row=range_.min_row, column=range_.min_col).value
        for col in range(range_.min_col, range_.max_col + 1):
            merged_headers[col] = top_left_cell_value

    # Obtenir les en-têtes pour les deux premières lignes, en supposant que la ligne 1 a les titres principaux
    for col in sheet.iter_cols(min_col=4, max_col=sheet.max_column, min_row=1, max_row=2):
        # Si la cellule n'est pas une partie d'une plage fusionnée
        if col[0].column not in merged_headers:
            headers.append(col[0].value)
        else:
            # Si la colonne est dans une plage fusionnée, vérifier si la deuxième ligne a une valeur
            # Sinon, utiliser uniquement la valeur de la première ligne (titre principal)
            secondary_header = col[1].value if col[1].value else ''
            main_header = merged_headers[col[0].column]
            header_title = f"{main_header} {secondary_header}".strip()
            headers.append(header_title)
    return headers


# Fonction pour extraire les hyperliens d'une cellule
def get_countries_responses(sheet, questions):
    # Dictionnaire pour stocker les réponses pour chaque pays
    responses_for_countries = {}
    for row in sheet.iter_rows():
        country = row[0].value
        urls = []
        for cell in row:
            if cell.hyperlink:  # verifye si la cellule a un lien hypertexte
                # hyperlink_texts.append((cell.value, cell.hyperlink.target))
                urls.append((cell.value, cell.hyperlink.target))

        # Parcourir chaque URL et poser les questions
        for url in urls:
            print(f"Processing document for {country} from URL: {url[1]}")
            responses = send_to_process_document(url[1], questions)
            if responses:
                # Ajouter les réponses dans le dictionnaire
                responses_for_countries[country] = responses
    return responses_for_countries


def write_countries_responses(sheet, countries_responses, questions):
    # Trouver l'index de chaque question dans les en-têtes
    question_indices = {question: questions.index(question) for question in questions}

    # Créer un dictionnaire pour suivre les lignes par pays (si nécessaire)
    country_row_map = {}
    current_row = 3  # Commencer à écrire à partir de la deuxième ligne

    # Parcourir les réponses pour chaque pays
    for country, data in countries_responses.items():
        if country not in country_row_map:
            # Si le pays n'est pas encore dans la feuille, ajoutez-le à une nouvelle ligne
            country_row_map[country] = current_row
            sheet.cell(row=current_row, column=1).value = country  # Mettre le nom du pays dans la première colonne
            current_row += 1  # Incrementer la ligne pour le prochain pays

        # Écrire les réponses pour le pays
        answers = data['answers']
        for answer in answers:
            question = answer['question']
            response = answer['answer']
            if question in question_indices:
                col_index = question_indices[
                                question] + 4  # +4 pour ajuster avec les colonnes (les questions commencent à la 4e colonne)
                sheet.cell(row=country_row_map[country], column=col_index).value = response


# Charger le fichier Excel
workbook = openpyxl.load_workbook(excel_file_path, data_only=True)

# Lire le contenu des feuilles 'Sources' et 'Data Collected'
sheet_sources = workbook['Sources']
sheet_data_collected = workbook['Data Collected']

# Préparer les questions en excluant les premières colonnes (Country, Cohort, Status)
questions = get_merged_cell_headers(sheet_data_collected)
# questions = sheet_data_collected.columns.tolist()[3:]

countries_responses = get_countries_responses(sheet_sources, questions)

# Écrire les réponses
write_countries_responses(sheet_data_collected, countries_responses, questions)

# Sauvegarder le fichier Excel
workbook.save('Updated_Data_Evidence.xlsx')

print(countries_responses)
